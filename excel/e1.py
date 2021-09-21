from openpyxl import Workbook

if __name__ == '__main__':
    # 创建workbook
    wb = Workbook()
    # 创建worksheet
    ws = wb.create_sheet('first_sheet')
    ws.title = '第一个工作簿'

    # 填充数据
    start_col = 'A'
    end_col = 'T'
    start_row = 1
    end_row = 11

    for i in range(ord(start_col), ord(end_col)):
        for j in range(start_row, end_row):
            cell_str = f'{chr(i)}{j}'
            ws[cell_str] = cell_str

    wb.worksheets.append(ws)
    wb.save("test.xlsx")